"""
Regenerates EIT_Multiplexer-BOM.xlsx from the canonical CSV + known manufacturer data.
Run from the BOM directory:  python3 build_bom.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# BOM data — source of truth is the KiCad CSV; manufacturer info hand-merged
# from the original Excel where designators matched.
# ---------------------------------------------------------------------------
ROWS = [
    # (Item, Qty, Reference, Value, Voltage, Size, Dielectric/Notes,
    #  Tolerance, Manufacturer, Manuf_Part, Description)

    # --- Resistors ---
    (1,  2, "R1, R5",                               "1kΩ",   "",    "0402", "",     "1%",   "Yageo",          "RC0402FR-071KP",        "RES SMD 1K OHM 1% 1/16W 0402"),
    (2,  2, "R2, R6",                               "10kΩ",  "",    "0402", "",     "1%",   "Yageo",          "RC0402FR-0710KP",       "RES SMD 10K OHM 1% 1/16W 0402"),
    (3,  2, "R3, R4",                               "27Ω",   "",    "0402", "",     "1%",   "Yageo",          "RC0402FR-0727RL",       "RES SMD 27 OHM 1% 1/16W 0402"),
    (4,  2, "R_INVERT1, R_INVERT2",                 "TBD",   "",    "0805", "",     "",     "",               "",                      "Invert resistors — value TBD"),

    # --- Capacitors ---
    (5,  2, "C2, C3",                               "15pF",  "50V", "0402", "C0G",  "5%",  "Murata",         "GCM1555C1H150JA16D",    "CAP CER 15PF 50V C0G/NP0 0402"),
    (6,  10,"C5,C6,C7,C9,C11,C12,C13,C14,C15,C16", "100nF", "25V", "0402", "X7R",  "10%", "Murata",         "GRM155R71E104KE14J",    "CAP CER 0.1UF 25V X7R 0402"),
    (7,  2, "C8, C10",                              "1µF",   "10V", "0402", "X6S",  "10%", "Murata",         "GRM155C81A105KA12D",    "CAP CER 1UF 10V X6S 0402"),
    (8,  2, "C1, C4",                               "10µF",  "16V", "0805", "X5R",  "10%", "Murata",         "GRM21BR61C106KE15K",    "CAP CER 10UF 16V X5R 0805"),
    (9,  8, "C17,C22,C23,C24,C25,C26,C27,C28",     "100nF", "25V", "0805", "X7R",  "10%", "",               "",                      "CAP CER 0.1UF 25V X7R 0805 — bypass for mux ICs"),
    (10, 2, "C18, C20",                             "4.7µF", "10V", "0805", "X5R",  "10%", "",               "",                      "CAP CER 4.7UF 10V X5R 0805"),
    (11, 1, "C19",                                  "1µF",   "16V", "0805", "X5R",  "10%", "",               "",                      "CAP CER 1UF 16V X5R 0805"),
    (12, 1, "C21",                                  "2.2µF", "10V", "0805", "X5R",  "10%", "",               "",                      "CAP CER 2.2UF 10V X5R 0805"),

    # --- ICs ---
    (13, 1, "U1",                                   "NCP1117-3.3", "","SOT-223-3","","",  "ON Semiconductor","NCP1117ST33T3G",        "IC REG LINEAR 3.3V 1A SOT223"),
    (14, 1, "U2",                                   "W25Q128JVS",  "","SOIC-8",   "","",  "Winbond",         "W25Q128JVSIQ TR",       "IC FLASH 128M SPI 133MHZ 8SOIC"),
    (15, 1, "U3",                                   "RP2040",      "","QFN-56",   "","",  "Raspberry Pi",    "RP2040",                "MCU ARM Cortex-M0+ dual-core 133MHz QFN56"),
    (16, 1, "U4",                                   "LM27761DSGR", "","DSG-8",    "","",  "Texas Instruments","LM27761DSGR",          "IC CONV DC/DC NEG 200MA 8WSON"),
    (17, 4, "U5, U6, U7, U8",                       "ADG1606BCPZ-REEL7","","CP-32-2","","","Analog Devices", "ADG1606BCPZ-REEL7",     "IC MUX 16:1 SINGLE 32LFCSP"),

    # --- Crystal ---
    (18, 1, "Y1",                                   "12MHz ABM8-272-T3","","SMD 3225","","","Abracon",        "ABM8-272-T3",           "CRYSTAL 12.0000MHZ 3.2x2.5MM SMD"),

    # --- Connectors ---
    (19, 1, "J1",                                   "USB Micro-B", "","","","",    "Amphenol",       "10103594-0001LF",       "CONN RCPT USB2.0 MICRO B SMD R/A"),
    (20, 1, "J2",                                   "1x16 pin socket","","2.54mm","","","","",              "Pin socket 1x16 2.54mm vertical"),

    # --- Switches ---
    (21, 2, "SW1, SW2",                             "Pushbutton",  "","","","",    "",               "",                      "Tactile push button SMD"),

    # --- Test Points ---
    (22, 8, "GND(x4), I+, I-, V+, V-",             "Test Point",  "","THT D2mm","","", "",              "",                      "THT test point pad D2.0mm drill 1.0mm"),

    # --- Mechanical ---
    (23, 4, "H1, H2, H3, H4",                      "Mounting Hole","","M2.5","","",  "",               "",                      "Mounting hole M2.5"),
]

HEADERS = ["Item", "Qty", "Reference", "Value", "Voltage", "Size",
           "Dielectric / Notes", "Tolerance", "Manufacturer",
           "Manuf. Part #", "Description"]

# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BOM"

# Title row
ws.merge_cells("A1:K1")
title_cell = ws["A1"]
title_cell.value = "EIT Multiplexer — Bill of Materials"
title_cell.font = Font(bold=True, size=13)
title_cell.alignment = Alignment(horizontal="center")

# Header row (row 2)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
thin = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col, h in enumerate(HEADERS, start=1):
    cell = ws.cell(row=2, column=col, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center")
    cell.border = border

# Section labels and row fills
SECTION_RANGES = {
    "Resistors":    range(1, 5),
    "Capacitors":   range(5, 13),
    "ICs":          range(13, 18),
    "Crystal":      range(18, 19),
    "Connectors":   range(19, 21),
    "Switches":     range(21, 22),
    "Test Points":  range(22, 23),
    "Mechanical":   range(23, 24),
}
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")  # light blue section header
ALT_FILL     = PatternFill("solid", fgColor="F2F7FB")  # alternating row

section_rows = {}
excel_row = 3
for section, item_range in SECTION_RANGES.items():
    # section header
    ws.merge_cells(f"A{excel_row}:K{excel_row}")
    sh = ws.cell(row=excel_row, column=1, value=section)
    sh.fill = SECTION_FILL
    sh.font = Font(bold=True, italic=True)
    excel_row += 1

    for item_idx in item_range:
        row_data = ROWS[item_idx - 1]
        fill = ALT_FILL if (excel_row % 2 == 0) else PatternFill()
        for col, val in enumerate(row_data, start=1):
            cell = ws.cell(row=excel_row, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="left" if col > 1 else "center",
                                       wrap_text=(col == 11))
            if fill.fill_type:
                cell.fill = fill
        excel_row += 1

# Column widths
COL_WIDTHS = [6, 5, 38, 18, 8, 12, 20, 10, 22, 26, 46]
for i, w in enumerate(COL_WIDTHS, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A3"

out = "/Users/erenerenel/Desktop/EIT-Multiplexer-PCB/Production/BOM/EIT_Multiplexer-BOM.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Total data rows: {excel_row - 3}")
